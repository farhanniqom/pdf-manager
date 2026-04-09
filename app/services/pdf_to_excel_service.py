"""Convert PDF files to Excel (.xlsx) by extracting tables with pdfplumber."""

from __future__ import annotations

import tempfile
import os
from io import BytesIO

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def _style_worksheet(ws, table_data: list[list]) -> None:
    """Apply clean formatting to a worksheet with extracted table data."""
    if not table_data:
        return

    # Header style
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(
        start_color="2B579A",
        end_color="2B579A",
        fill_type="solid")
    header_alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True)

    # Body style
    body_font = Font(name="Calibri", size=11)
    body_alignment = Alignment(vertical="center", wrap_text=True)

    # Border style
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    alt_fill = PatternFill(
        start_color="F2F7FB",
        end_color="F2F7FB",
        fill_type="solid")

    for row_idx, row in enumerate(table_data, start=1):
        for col_idx, cell_value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
            cell.border = thin_border
            if row_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            else:
                cell.font = body_font
                cell.alignment = body_alignment
                if row_idx % 2 == 0:
                    cell.fill = alt_fill

    # Auto-fit column widths
    for col_idx in range(1, len(table_data[0]) + 1 if table_data else 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for row in table_data:
            if col_idx - 1 < len(row):
                cell_val = str(row[col_idx - 1]
                               ) if row[col_idx - 1] is not None else ""
                max_length = max(max_length, len(cell_val))
        adjusted_width = min(max(max_length + 4, 10), 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def convert_pdf_to_excel(data: bytes) -> bytes:
    """
    Convert PDF bytes (already decrypted if needed) to XLSX bytes.

    Extracts tables from each page using pdfplumber.
    Each page with tables gets its own sheet in the workbook.
    Tables are formatted with headers, borders, and alternating row colors.

    Returns the XLSX content as bytes.
    Raises RuntimeError on conversion failure.
    """
    tmp_dir = tempfile.mkdtemp()
    pdf_path = os.path.join(tmp_dir, "input.pdf")

    try:
        with open(pdf_path, "wb") as f:
            f.write(data)

        wb = Workbook()
        # Remove the default sheet; we'll create named ones
        wb.remove(wb.active)

        sheet_count = 0

        with pdfplumber.open(pdf_path) as pdf:
            for page_num, page in enumerate(pdf.pages, start=1):
                tables = page.extract_tables()

                if tables:
                    for table_idx, table in enumerate(tables):
                        if not table:
                            continue

                        sheet_count += 1
                        if len(tables) > 1:
                            sheet_name = f"Halaman {page_num} - Tabel {table_idx + 1}"
                        else:
                            sheet_name = f"Halaman {page_num}"

                        # Excel sheet name max 31 chars
                        sheet_name = sheet_name[:31]

                        ws = wb.create_sheet(title=sheet_name)

                        # Clean up cell values
                        cleaned_table = []
                        for row in table:
                            cleaned_row = []
                            for cell in row:
                                if cell is None:
                                    cleaned_row.append("")
                                else:
                                    cleaned_row.append(
                                        str(cell).strip().replace("\n", " "))
                            cleaned_table.append(cleaned_row)

                        _style_worksheet(ws, cleaned_table)

        # If no tables were found on any page, try extracting text as fallback
        if sheet_count == 0:
            with pdfplumber.open(pdf_path) as pdf:
                ws = wb.create_sheet(title="Konten PDF")
                row_num = 1
                has_content = False

                for page_num, page in enumerate(pdf.pages, start=1):
                    text = page.extract_text()
                    if text:
                        has_content = True
                        # Add page header
                        header_cell = ws.cell(
                            row=row_num, column=1,
                            value=f"--- Halaman {page_num} ---",
                        )
                        header_cell.font = Font(
                            name="Calibri", bold=True, size=12, color="2B579A",
                        )
                        row_num += 1

                        for line in text.split("\n"):
                            stripped = line.strip()
                            if stripped:
                                ws.cell(row=row_num, column=1, value=stripped)
                                row_num += 1
                        row_num += 1  # blank row between pages

                ws.column_dimensions["A"].width = 80

                if not has_content:
                    raise RuntimeError(
                        "Tidak ditemukan tabel atau teks yang dapat diekstrak dari PDF."
                    )

        # Write workbook to bytes
        output = BytesIO()
        wb.save(output)
        result = output.getvalue()

        if not result:
            raise RuntimeError("Konversi menghasilkan file kosong.")

        return result

    except RuntimeError:
        raise
    except Exception as e:
        raise RuntimeError(f"Gagal mengonversi PDF ke Excel: {e}") from e
    finally:
        for p in (pdf_path,):
            try:
                os.unlink(p)
            except OSError:
                pass
        try:
            os.rmdir(tmp_dir)
        except OSError:
            pass
